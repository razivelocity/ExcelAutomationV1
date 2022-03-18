
import openpyxl as xl
import pandas as pd
#Loading of teh Excel
wb = xl.load_workbook('C:\\Users\\DELL\\OneDrive\\Razi_Rahman\\Desktop\\CityTemperature.xlsx')
sheet = wb['Sheet1']
    # create new sheet
    #wb.create_sheet('Sheet 2')
    #Sheet2=wb['Sheet 2']

#Looping through Excel
for row in range(2, sheet.max_row + 1):
        ColumnA_Data= sheet.cell(row, 1).value
        ColumnB_Data= sheet.cell(row,2).value
        ColumnC_Data= sheet.cell(row,3).value
        print(ColumnC_Data)
        print(ColumnB_Data)
        print(ColumnA_Data)
        sheet['A8'].value="Checking"
        sheet['A'+str(9)].value="good"

#Conversion of Excel into CSV
    # Read and store content
    # of an excel file
read_file = pd.read_excel ("C:\\Users\\DELL\\OneDrive\\Razi_Rahman\\Desktop\\CityTemperature.xlsx")

    # Write the dataframe object
    # into csv file
read_file.to_csv ("C:\\Users\\DELL\\OneDrive\\Razi_Rahman\\Desktop\\CityTemperature_csv.csv",
                      index = None,
                      header=True)


    #Insertion of Columns and deletion of Column

sheet.insert_cols(2)
sheet.delete_cols(2)

wb.save('C:\\Users\\DELL\\OneDrive\\Razi_Rahman\\Desktop\\CityTemperature.xlsx')


